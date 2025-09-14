#!/usr/bin/env python3
"""
Модуль для работы с Excel файлами
Создает и обновляет Excel файлы с заявками на рейсы
"""

import os
import logging
from datetime import datetime
from typing import List, Optional

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

MASTER_EXCEL_FILENAME = "все_заявки.xlsx"
ALL_OKES = ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    # Fallback для случаев, когда openpyxl не установлен
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl не установлен. Excel функциональность будет недоступна.")

class ExcelIntegration:
    """Класс для работы с Excel файлами заявок на рейсы"""
    
    def __init__(self, data_dir: str = "data"):
        """
        Инициализация Excel интеграции
        
        Args:
            data_dir: Директория для хранения Excel файлов
        """
        self.data_dir = data_dir
        self.ensure_data_dir()
        
        # Заголовки для листов
        self.headers = [
            "Время подачи заявки",
            "ФИО", 
            "Табельный номер",
            "Должность",
            "Направление",
            "Информация о рейсе"
        ]
        
        logger.info(f"📊 Excel интеграция инициализирована. Директория: {self.data_dir}")
        
        self.master_file_path = os.path.join(self.data_dir, MASTER_EXCEL_FILENAME)
        self.workbook = None # Будет инициализирован в _initialize_master_excel_file
        self._initialize_master_excel_file()

    def _initialize_master_excel_file(self):
        """Инициализирует или загружает мастер-файл Excel и создает все листы ОКЭ"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl не установлен. Мастер-файл Excel не будет инициализирован.")
            return

        if os.path.exists(self.master_file_path):
            self.workbook = load_workbook(self.master_file_path)
            logger.info(f"Загружен существующий мастер-файл Excel: {self.master_file_path}")
        else:
            self.workbook = Workbook()
            # Удаляем лист по умолчанию, который создается автоматически, если он есть
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            logger.info(f"Создан новый мастер-файл Excel: {self.master_file_path}")

        for oke_name in ALL_OKES:
            if oke_name not in self.workbook.sheetnames:
                ws = self.workbook.create_sheet(title=oke_name)
                # Добавляем заголовки
                for col, header in enumerate(self.headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Автоподбор ширины колонок (начальный)
                for col_idx, header in enumerate(self.headers, 1):
                    adjusted_width = min(len(header) + 2, 50)
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

                logger.info(f"Создан лист '{oke_name}' в мастер-файле.")
        self.workbook.save(self.master_file_path)

    def ensure_data_dir(self):
        """Создает директорию для данных, если она не существует"""
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            logger.info(f"📁 Создана директория: {self.data_dir}")
    
    def get_excel_file_path(self, oke: str) -> str:
        """
        Получает путь к Excel файлу для указанного ОКЭ
        
        Args:
            oke: Название ОКЭ (например, "ОКЭ 1")
            
        Returns:
            Путь к Excel файлу
        """
        # Очищаем название ОКЭ от недопустимых символов для имени файла
        safe_oke = "".join(c for c in oke if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_oke = safe_oke.replace(' ', '_')
        
        filename = f"{safe_oke}_заявки.xlsx"
        return os.path.join(self.data_dir, filename)
    
    def create_excel_file(self, oke: str) -> bool:
        """
        Создает новый Excel файл с заголовками для указанного ОКЭ
        
        Args:
            oke: Название ОКЭ
            
        Returns:
            True если файл создан успешно, False в противном случае
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl не установлен. Невозможно создать Excel файл.")
            return False
            
        try:
            file_path = self.get_excel_file_path(oke)
            
            # Создаем новую рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Заявки на рейсы"
            
            # Добавляем заголовки
            for col, header in enumerate(self.headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                # Стилизация заголовков
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            wb.save(file_path)
            logger.info(f"✅ Создан Excel файл: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания Excel файла для {oke}: {e}")
            return False
    
    def file_exists(self, oke: str) -> bool:
        """
        Проверяет, существует ли лист для указанного ОКЭ в мастер-файле Excel
        
        Args:
            oke: Название ОКЭ
            
        Returns:
            True если лист существует, False в противном случае
        """
        if not OPENPYXL_AVAILABLE or self.workbook is None:
            return False
        return oke in self.workbook.sheetnames
    
    def add_application(self, oke: str, fio: str, tab_num: str, position: str, 
                       direction: str, flight_info: str) -> bool:
        """
        Добавляет новую заявку в соответствующий лист мастер-файла Excel
        
        Args:
            oke: Название ОКЭ
            fio: ФИО сотрудника
            tab_num: Табельный номер
            position: Должность
            direction: Направление
            flight_info: Информация о рейсе
            
        Returns:
            True если заявка добавлена успешно, False в противном случае
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl не установлен. Невозможно добавить заявку в Excel.")
            return False
            
        try:
            # Загружаем существующую рабочую книгу
            wb = load_workbook(self.master_file_path)
            ws = wb[oke]
            
            # Находим следующую свободную строку
            next_row = ws.max_row + 1
            
            # Подготавливаем данные для записи
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            data = [timestamp, fio, tab_num, position, direction, flight_info]
            
            # Записываем данные
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=next_row, column=col, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Автоподбор ширины колонок для новых данных
            for col_idx, value in enumerate(data, 1):
                column_letter = ws.cell(row=next_row, column=col_idx).column_letter
                max_length = max((len(str(cell.value)) for cell in ws[column_letter] if cell.value is not None), default=0)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            wb.save(self.master_file_path)
            logger.info(f"✅ Заявка добавлена в {oke} (Excel): {fio} - {direction}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Ошибка добавления заявки в {oke} (Excel): {e}")
            return False
    
    def get_applications_count(self, oke: str) -> int:
        """
        Получает количество заявок в Excel файле для указанного ОКЭ
        
        Args:
            oke: Название ОКЭ
            
        Returns:
            Количество заявок (без учета заголовка)
        """
        if not OPENPYXL_AVAILABLE or self.workbook is None:
            logger.error("openpyxl не установлен. Невозможно получить количество заявок.")
            return 0
            
        try:
            # Перезагружаем рабочую книгу перед подсчетом, чтобы учесть последние изменения
            self.workbook = load_workbook(self.master_file_path)

            if oke not in self.workbook.sheetnames:
                return 0
                
            ws = self.workbook[oke]
            
            # Количество строк минус заголовок
            return max(0, ws.max_row - 1)
            
        except Exception as e:
            logger.error(f"❌ Ошибка подсчета заявок для {oke} (Excel): {e}")
            return 0
    
    def get_all_files(self) -> List[str]:
        """
        Получает список всех Excel файлов с заявками
        
        Returns:
            Список путей к Excel файлам
        """
        # Этот метод возвращает путь к мастер-файлу, если openpyxl доступен
        if not OPENPYXL_AVAILABLE or self.workbook is None or not os.path.exists(self.master_file_path):
            return []
        return [self.master_file_path]
    
    def read_applications(self, oke: str) -> List[dict]:
        """
        Читает все заявки из указанного ОКЭ
        
        Args:
            oke: Название ОКЭ
            
        Returns:
            Список словарей с данными заявок
        """
        try:
            if not OPENPYXL_AVAILABLE or self.workbook is None or not os.path.exists(self.master_file_path):
                return []
            
            if oke not in self.workbook.sheetnames:
                return []
            
            ws = self.workbook[oke]
            applications = []
            
            # Читаем данные начиная со второй строки (первая - заголовки)
            for row in range(2, ws.max_row + 1):
                app_data = {}
                for col, header in enumerate(self.headers, 1):
                    cell_value = ws.cell(row=row, column=col).value
                    app_data[header] = str(cell_value) if cell_value is not None else ""
                
                # Добавляем только если есть хотя бы ФИО
                if app_data.get('ФИО', '').strip():
                    applications.append(app_data)
            
            logger.info(f"📖 Прочитано {len(applications)} заявок из {oke}")
            return applications
            
        except Exception as e:
            logger.error(f"❌ Ошибка чтения заявок из {oke} (Excel): {e}")
            return []
    
    def get_statistics(self) -> dict:
        """
        Получает статистику по всем заявкам
        
        Returns:
            Словарь со статистикой
        """
        try:
            stats = {
                "total_files": 0,
                "total_applications": 0,
                "by_oke": {}
            }
            
            if not OPENPYXL_AVAILABLE or self.workbook is None or not os.path.exists(self.master_file_path):
                return stats

            stats["total_files"] = 1  # Один мастер-файл
            
            for oke_name in ALL_OKES:
                if oke_name in self.workbook.sheetnames:
                    count = self.get_applications_count(oke_name)
                    stats["by_oke"][oke_name] = count
                    stats["total_applications"] += count
            
            return stats
            
        except Exception as e:
            logger.error(f"❌ Ошибка получения статистики (Excel): {e}")
            return {"total_files": 0, "total_applications": 0, "by_oke": {}}


def test_excel_integration():
    """Тестовая функция для проверки работы Excel интеграции (новый формат)"""
    print("🧪 Тестирование Excel интеграции (новый формат)...")
    
    excel = ExcelIntegration()
    
    # Удаляем старый мастер-файл для чистого теста
    if os.path.exists(excel.master_file_path):
        os.remove(excel.master_file_path)
        print(f"🗑️ Удален старый мастер-файл: {excel.master_file_path}")
    
    excel = ExcelIntegration() # Переинициализируем после удаления

    # Тест добавления заявки в ОКЭ 1
    print("📋 Добавление тестовой заявки в ОКЭ 1...")
    if excel.add_application(
        oke="ОКЭ 1",
        fio="Тестов Те.Т.",
        tab_num="123456",
        position="БП",
        direction="Москва",
        flight_info="15.09.2024 Утренний рейс"
    ):
        print("✅ Заявка добавлена в ОКЭ 1 успешно")
    else:
        print("❌ Ошибка добавления заявки в ОКЭ 1")
        return

    # Тест добавления заявки в ОЛСиТ
    print("📋 Добавление тестовой заявки в ОЛСиТ...")
    if excel.add_application(
        oke="ОЛСиТ",
        fio="Тестова А.А.",
        tab_num="654321",
        position="СБЭ",
        direction="Санкт-Петербург",
        flight_info="20.10.2024 Вечерний рейс"
    ):
        print("✅ Заявка добавлена в ОЛСиТ успешно")
    else:
        print("❌ Ошибка добавления заявки в ОЛСиТ")
        return
    
    # Тест статистики
    print("📊 Получение статистики...")
    stats = excel.get_statistics()
    print(f"📈 Статистика: {stats}")
    
    # Проверка количества заявок в листах
    if stats["by_oke"].get("ОКЭ 1") == 1 and stats["by_oke"].get("ОЛСиТ") == 1:
        print("✅ Количество заявок в листах верное")
    else:
        print("❌ Неверное количество заявок в листах")

    print("🎉 Тест завершен успешно!")


if __name__ == "__main__":
    test_excel_integration()
