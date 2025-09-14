#!/usr/bin/env python3
"""
Express SmartApp - Flight Booking Bot
Веб-приложение для заказа рейсов в Express SmartApp
"""

import os
import time
import logging
from datetime import datetime, timedelta
from dataclasses import dataclass
from enum import Enum
from typing import Dict, List, Optional, Any
from threading import Lock

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory
from flask_cors import CORS
from excel_integration import ExcelIntegration
from user_management import user_manager, UserRole, UserStatus
from notification_system import notification_manager, NotificationType, NotificationStatus
from express_integration import init_express_integration, get_express_bot, get_notification_service

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Конфигурация
@dataclass
class SmartAppConfig:
    """Конфигурация SmartApp"""
    app_name: str = "Flight Booking Bot"
    app_version: str = "1.0.0"
    secret_key: str = os.getenv('SMARTAPP_SECRET_KEY', 'your-secret-key-here')
    debug: bool = os.getenv('DEBUG', 'False').lower() == 'true'

# Состояния пользователя
class UserState(Enum):
    """Состояния пользователя в SmartApp"""
    LOCATION_SELECTION = "location_selection"
    OKE_SELECTION = "oke_selection"
    DATE_SELECTION = "date_selection"
    POSITION_SELECTION = "position_selection"
    FIO_INPUT = "fio_input"
    DIRECTION_SELECTION = "direction_selection"
    FLIGHT_WISHES = "flight_wishes"
    CONFIRMATION = "confirmation"
    COMPLETED = "completed"

# Данные пользователя
@dataclass
class UserData:
    """Данные пользователя для заявки"""
    user_id: str
    location: Optional[str] = None
    oke: Optional[str] = None
    selected_date: Optional[str] = None
    position: Optional[str] = None
    fio: Optional[str] = None
    tab_num: Optional[str] = None
    direction: Optional[str] = None
    flight_wishes: Optional[str] = None
    state: UserState = UserState.LOCATION_SELECTION
    created_at: datetime = None
    
    def __post_init__(self):
        if self.created_at is None:
            self.created_at = datetime.now()

# Словарь направлений по локациям
DIRECTIONS = {
    'МСК': [
        'Анадырь', 'Благовещенск', 'Владивосток', 'Красноярск', 'Магадан',
        'Нижний Новгород', 'Самара', 'Санкт-Петербург', 'Уфа', 'Хабаровск',
        'Южно-Сахалинск', 'Варадеро', 'Хургада', 'Шарм-Эль-Шейх'
    ],
    'СПБ': [
        'Волгоград', 'Красноярск', 'Москва', 'Самара', 'Сочи',
        'Хургада', 'Шарм-Эль-Шейх'
    ],
    'Красноярск': [
        'Владивосток', 'Сочи', 'Южно-Сахалинск', 'Хабаровск', 'Санья'
    ],
    'Сочи': []  # Для Сочи пока нет предопределенных направлений
}

# Создание Flask приложения
app = Flask(__name__)
app.config['SECRET_KEY'] = SmartAppConfig().secret_key
CORS(app, origins=['http://localhost:8080', 'http://127.0.0.1:8080', 'https://thinking-britain-leu-browsing.trycloudflare.com'], supports_credentials=True)  # Разрешить CORS для всех доменов

# Глобальные переменные
config = SmartAppConfig()
user_data: Dict[str, UserData] = {}
data_lock = Lock()

# Инициализация простого логгера (Удаляем)
excel_integration = ExcelIntegration(data_dir="data")

# Инициализация Express интеграции
EXPRESS_BOT_TOKEN = os.getenv('EXPRESS_BOT_TOKEN', 'your_bot_token_here')
EXPRESS_WEBHOOK_SECRET = os.getenv('EXPRESS_WEBHOOK_SECRET', 'your_webhook_secret_here')
init_express_integration(EXPRESS_BOT_TOKEN, EXPRESS_WEBHOOK_SECRET)

def get_user_data(user_id: str) -> UserData:
    """Получить данные пользователя"""
    with data_lock:
        if user_id not in user_data:
            user_data[user_id] = UserData(user_id=user_id)
        return user_data[user_id]

def save_user_data(user_data_obj: UserData) -> None:
    """Сохранить данные пользователя"""
    with data_lock:
        user_data[user_data_obj.user_id] = user_data_obj

def clear_user_data(user_id: str) -> None:
    """Очистить данные пользователя"""
    with data_lock:
        if user_id in user_data:
            del user_data[user_id]

def generate_calendar_days(year: int, month: int) -> List[Dict[str, Any]]:
    """Генерировать дни календаря"""
    today = datetime.now().date()
    first_day = datetime(year, month, 1).date()
    last_day = (datetime(year, month + 1, 1) - timedelta(days=1)).date()
    
    days = []
    current_date = first_day
    
    while current_date <= last_day:
        is_today = current_date == today
        is_past = current_date < today
        
        days.append({
            'day': current_date.day,
            'date': current_date.strftime('%d.%m.%Y'),
            'is_today': is_today,
            'is_past': is_past,
            'disabled': is_past
        })
        current_date += timedelta(days=1)
    
    return days

def get_month_name(month: int) -> str:
    """Получить название месяца"""
    months = [
        'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
        'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
    ]
    return months[month - 1]

@app.route('/')
def index():
    """Главная страница SmartApp"""
    user_id = request.args.get('user_id', 'default_user')
    user_data_obj = get_user_data(user_id)
    
    return render_template('smartapp_index.html',
                         user_data=user_data_obj,
                         config=config)

@app.route('/flight_booking_ui.html')
def flight_booking_ui():
    """Страница заявки на рейс (десктоп)"""
    return send_from_directory('.', 'flight_booking_ui.html')

@app.route('/mobile_booking_ui.html')
def mobile_booking_ui():
    """Страница заявки на рейс (мобильная)"""
    return send_from_directory('.', 'mobile_booking_ui.html')

@app.route('/api/start', methods=['POST'])
def api_start():
    """API для начала процесса заявки"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        
        # Очищаем предыдущие данные
        clear_user_data(user_id)
        user_data_obj = UserData(user_id=user_id)
        save_user_data(user_data_obj)
        
        return jsonify({
            'success': True,
            'message': 'Процесс заявки начат',
            'state': user_data_obj.state.value
        })
    except Exception as e:
        logger.error(f"Ошибка в api_start: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/select-location', methods=['POST'])
def api_select_location():
    """API для выбора локации"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        location = data.get('location')
        
        if not location:
            return jsonify({'success': False, 'error': 'Локация не указана'}), 400
        
        user_data_obj = get_user_data(user_id)
        user_data_obj.location = location
        user_data_obj.state = UserState.OKE_SELECTION
        save_user_data(user_data_obj)
        
        return jsonify({
            'success': True,
            'message': f'Выбрана локация: {location}',
            'state': user_data_obj.state.value
        })
    except Exception as e:
        logger.error(f"Ошибка в api_select_location: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/select-oke', methods=['POST'])
def api_select_oke():
    """API для выбора ОКЭ"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        oke = data.get('oke')
        
        if not oke:
            return jsonify({'success': False, 'error': 'ОКЭ не указано'}), 400
        
        user_data_obj = get_user_data(user_id)
        user_data_obj.oke = oke
        user_data_obj.state = UserState.DATE_SELECTION
        save_user_data(user_data_obj)
        
        return jsonify({
            'success': True,
            'message': f'Выбрано ОКЭ: {oke}',
            'state': user_data_obj.state.value
        })
    except Exception as e:
        logger.error(f"Ошибка в api_select_oke: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/select-date', methods=['POST'])
def api_select_date():
    """API для выбора даты"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        selected_date = data.get('date')
        
        if not selected_date:
            return jsonify({'success': False, 'error': 'Дата не указана'}), 400
        
        user_data_obj = get_user_data(user_id)
        user_data_obj.selected_date = selected_date
        user_data_obj.state = UserState.POSITION_SELECTION
        save_user_data(user_data_obj)
        
        return jsonify({
            'success': True,
            'message': f'Выбрана дата: {selected_date}',
            'state': user_data_obj.state.value
        })
    except Exception as e:
        logger.error(f"Ошибка в api_select_date: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/select-position', methods=['POST'])
def api_select_position():
    """API для выбора должности"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        position = data.get('position')
        
        if not position:
            return jsonify({'success': False, 'error': 'Должность не указана'}), 400
        
        user_data_obj = get_user_data(user_id)
        user_data_obj.position = position
        user_data_obj.state = UserState.FIO_INPUT
        save_user_data(user_data_obj)
        
        return jsonify({
            'success': True,
            'message': f'Выбрана должность: {position}',
            'state': user_data_obj.state.value
        })
    except Exception as e:
        logger.error(f"Ошибка в api_select_position: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/input-fio', methods=['POST'])
def api_input_fio():
    """API для ввода ФИО и табельного номера"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        fio = data.get('fio')
        tab_num = data.get('tab_num')
        
        if not fio or not tab_num:
            return jsonify({'success': False, 'error': 'ФИО и табельный номер обязательны'}), 400
        
        user_data_obj = get_user_data(user_id)
        user_data_obj.fio = fio
        user_data_obj.tab_num = tab_num
        user_data_obj.state = UserState.DIRECTION_SELECTION
        save_user_data(user_data_obj)
        
        return jsonify({
            'success': True,
            'message': 'ФИО и табельный номер сохранены',
            'state': user_data_obj.state.value
        })
    except Exception as e:
        logger.error(f"Ошибка в api_input_fio: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/select-direction', methods=['POST'])
def api_select_direction():
    """API для выбора направления"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        direction = data.get('direction')
        
        if not direction:
            return jsonify({'success': False, 'error': 'Направление не указано'}), 400
        
        user_data_obj = get_user_data(user_id)
        user_data_obj.direction = direction
        user_data_obj.state = UserState.FLIGHT_WISHES
        save_user_data(user_data_obj)
        
        return jsonify({
            'success': True,
            'message': f'Выбрано направление: {direction}',
            'state': user_data_obj.state.value
        })
    except Exception as e:
        logger.error(f"Ошибка в api_select_direction: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/input-wishes', methods=['POST'])
def api_input_wishes():
    """API для ввода пожеланий"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        wishes = data.get('wishes', '-')
        
        user_data_obj = get_user_data(user_id)
        user_data_obj.flight_wishes = wishes
        user_data_obj.state = UserState.CONFIRMATION
        save_user_data(user_data_obj)
        
        return jsonify({
            'success': True,
            'message': 'Пожелания сохранены',
            'state': user_data_obj.state.value
        })
    except Exception as e:
        logger.error(f"Ошибка в api_input_wishes: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/application', methods=['POST', 'OPTIONS'])
def api_create_application():
    """API для создания заявки напрямую"""
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        data = request.get_json()
        logger.info(f"📝 Получены данные заявки: {data}")
        
        # Проверяем обязательные поля
        required_fields = ['location', 'oke', 'date', 'position', 'fio', 'tab_num', 'direction']
        missing_fields = []
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            error_msg = f'Отсутствуют обязательные поля: {", ".join(missing_fields)}'
            logger.error(f"❌ {error_msg}")
            return jsonify({'success': False, 'error': error_msg}), 400
        
        # Добавляем заявку в Excel
        if not excel_integration.add_application(
            oke=data['oke'],
            fio=data['fio'],
            tab_num=data['tab_num'],
            position=data['position'],
            direction=data['direction'],
            flight_info=f"{data['date']} {data.get('wishes', '')}"
        ):
            logger.error(f"❌ Не удалось сохранить заявку в Excel для '{data['oke']}'")
            return jsonify({'success': False, 'error': 'Не удалось сохранить заявку в Excel'}), 500
        
        logger.info(f"✅ Заявка успешно создана: {data['fio']} - {data['oke']}")
        return jsonify({
            'success': True,
            'message': 'Заявка успешно отправлена!',
            'application_id': f"app_{int(time.time())}"
        })
        
    except Exception as e:
        logger.error(f"❌ Ошибка создания заявки: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/confirm', methods=['POST'])
def api_confirm():
    """API для подтверждения заявки"""
    try:
        data = request.get_json()
        user_id = data.get('user_id', 'default_user')
        
        user_data_obj = get_user_data(user_id)
        
        # Проверяем, что все данные собраны
        if not all([user_data_obj.location, user_data_obj.oke, user_data_obj.selected_date,
                   user_data_obj.position, user_data_obj.fio, user_data_obj.tab_num,
                   user_data_obj.direction, user_data_obj.flight_wishes]):
            return jsonify({'success': False, 'error': 'Не все данные собраны'}), 400
        
        # Записываем заявку в лог
        try:
            full_flight_info = f"{user_data_obj.selected_date} {user_data_obj.flight_wishes}"

            # Добавляем заявку в Excel
            if not excel_integration.add_application(
                oke=user_data_obj.oke,
                fio=user_data_obj.fio,
                tab_num=user_data_obj.tab_num,
                position=user_data_obj.position,
                direction=user_data_obj.direction,
                flight_info=full_flight_info
            ):
                logger.error(f"❌ Не удалось сохранить заявку в Excel для '{user_data_obj.oke}'")
                return jsonify({'success': False, 'error': f'Не удалось сохранить заявку в Excel'}), 500
            
            user_data_obj.state = UserState.COMPLETED
            save_user_data(user_data_obj)
            
            return jsonify({
                'success': True,
                'message': 'Заявка успешно отправлена!',
                'state': user_data_obj.state.value
            })
        except Exception as e:
            logger.error(f"Ошибка при записи заявки: {e}")
            return jsonify({'success': False, 'error': f'Ошибка при сохранении: {str(e)}'}), 500
            
    except Exception as e:
        logger.error(f"Ошибка в api_confirm: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/calendar/<int:year>/<int:month>')
def api_calendar(year: int, month: int):
    """API для получения календаря"""
    try:
        days = generate_calendar_days(year, month)
        month_name = get_month_name(month)
        
        return jsonify({
            'success': True,
            'year': year,
            'month': month,
            'month_name': month_name,
            'days': days
        })
    except Exception as e:
        logger.error(f"Ошибка в api_calendar: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/directions/<location>')
def api_directions(location: str):
    """API для получения направлений по локации"""
    try:
        directions = DIRECTIONS.get(location, [])
        return jsonify({
            'success': True,
            'location': location,
            'directions': directions
        })
    except Exception as e:
        logger.error(f"Ошибка в api_directions: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/user-data/<user_id>')
def api_user_data(user_id: str):
    """API для получения данных пользователя"""
    try:
        user_data_obj = get_user_data(user_id)
        
        return jsonify({
            'success': True,
            'user_data': {
                'user_id': user_data_obj.user_id,
                'location': user_data_obj.location,
                'oke': user_data_obj.oke,
                'selected_date': user_data_obj.selected_date,
                'position': user_data_obj.position,
                'fio': user_data_obj.fio,
                'tab_num': user_data_obj.tab_num,
                'direction': user_data_obj.direction,
                'flight_wishes': user_data_obj.flight_wishes,
                'state': user_data_obj.state.value,
                'created_at': user_data_obj.created_at.isoformat()
            }
        })
    except Exception as e:
        logger.error(f"Ошибка в api_user_data: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/health')
def health():
    """Health check endpoint"""
    return jsonify({
        'status': 'ok',
        'app_name': config.app_name,
        'version': config.app_version,
        'active_users': len(user_data),
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/statistics', methods=['GET'])
def api_statistics():
    """API для получения статистики заявок"""
    try:
        stats = excel_integration.get_statistics() # Используем ExcelIntegration для статистики
        return jsonify({
            'success': True,
            'statistics': stats
        })
    except Exception as e:
        logger.error(f"Ошибка получения статистики: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/applications/<oke>', methods=['GET'])
def api_get_applications(oke):
    """API для получения заявок по ОКЭ"""
    try:
        applications = excel_integration.read_applications(oke) # Используем ExcelIntegration для получения заявок
        return jsonify({
            'success': True,
            'oke': oke,
            'applications': applications,
            'count': len(applications)
        })
    except Exception as e:
        logger.error(f"Ошибка получения заявок для {oke}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/search', methods=['POST'])
def api_search_applications():
    """API для поиска заявок с фильтрами"""
    try:
        data = request.get_json()
        
        # Параметры поиска
        search_query = data.get('query', '').strip()
        oke_filter = data.get('oke', '')
        date_from = data.get('date_from', '')
        date_to = data.get('date_to', '')
        position_filter = data.get('position', '')
        page = int(data.get('page', 1))
        per_page = int(data.get('per_page', 20))
        
        # Получаем все заявки из всех ОКЭ
        all_applications = []
        for oke in ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']:
            try:
                applications = excel_integration.read_applications(oke)
                for app in applications:
                    app['oke'] = oke  # Добавляем информацию об ОКЭ
                    all_applications.append(app)
            except Exception as e:
                logger.warning(f"Ошибка чтения заявок из {oke}: {e}")
                continue
        
        # Применяем фильтры
        filtered_applications = []
        for app in all_applications:
            # Фильтр по поисковому запросу
            if search_query:
                search_fields = [
                    app.get('ФИО', ''),
                    app.get('Табельный номер', ''),
                    app.get('Должность', ''),
                    app.get('Направление', ''),
                    app.get('Информация о рейсе', '')
                ]
                if not any(search_query.lower() in str(field).lower() for field in search_fields):
                    continue
            
            # Фильтр по ОКЭ
            if oke_filter and app.get('oke', '') != oke_filter:
                continue
            
            # Фильтр по должности
            if position_filter and position_filter.lower() not in app.get('Должность', '').lower():
                continue
            
            # Фильтр по дате
            if date_from or date_to:
                app_date_str = app.get('Время подачи заявки', '')
                try:
                    app_date = datetime.strptime(app_date_str.split()[0], '%Y-%m-%d')
                    if date_from:
                        from_date = datetime.strptime(date_from, '%Y-%m-%d')
                        if app_date < from_date:
                            continue
                    if date_to:
                        to_date = datetime.strptime(date_to, '%Y-%m-%d')
                        if app_date > to_date:
                            continue
                except (ValueError, IndexError):
                    continue
            
            filtered_applications.append(app)
        
        # Сортировка по дате (новые сначала)
        filtered_applications.sort(key=lambda x: x.get('Время подачи заявки', ''), reverse=True)
        
        # Пагинация
        total_count = len(filtered_applications)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_applications = filtered_applications[start_idx:end_idx]
        
        return jsonify({
            'success': True,
            'applications': paginated_applications,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_count': total_count,
                'total_pages': (total_count + per_page - 1) // per_page
            },
            'filters': {
                'query': search_query,
                'oke': oke_filter,
                'date_from': date_from,
                'date_to': date_to,
                'position': position_filter
            }
        })
        
    except Exception as e:
        logger.error(f"Ошибка поиска заявок: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/filters/options', methods=['GET'])
def api_get_filter_options():
    """API для получения опций фильтров"""
    try:
        # Получаем уникальные должности
        positions = set()
        oke_counts = {}
        
        for oke in ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']:
            try:
                applications = excel_integration.read_applications(oke)
                oke_counts[oke] = len(applications)
                
                for app in applications:
                    position = app.get('Должность', '').strip()
                    if position:
                        positions.add(position)
            except Exception as e:
                logger.warning(f"Ошибка чтения заявок из {oke}: {e}")
                oke_counts[oke] = 0
        
        return jsonify({
            'success': True,
            'options': {
                'okes': [{'name': oke, 'count': oke_counts.get(oke, 0)} for oke in ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']],
                'positions': sorted(list(positions)),
                'total_applications': sum(oke_counts.values())
            }
        })
        
    except Exception as e:
        logger.error(f"Ошибка получения опций фильтров: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export', methods=['POST'])
def api_export_applications():
    """API для экспорта заявок в Excel"""
    try:
        data = request.get_json()
        
        # Параметры экспорта
        oke_filter = data.get('oke', '')
        date_from = data.get('date_from', '')
        date_to = data.get('date_to', '')
        format_type = data.get('format', 'excel')  # excel, csv
        
        # Получаем заявки с фильтрами (используем логику из поиска)
        all_applications = []
        for oke in ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']:
            if oke_filter and oke != oke_filter:
                continue
                
            try:
                applications = excel_integration.read_applications(oke)
                for app in applications:
                    app['oke'] = oke
                    all_applications.append(app)
            except Exception as e:
                logger.warning(f"Ошибка чтения заявок из {oke}: {e}")
                continue
        
        # Применяем фильтры по дате
        if date_from or date_to:
            filtered_applications = []
            for app in all_applications:
                app_date_str = app.get('Время подачи заявки', '')
                try:
                    app_date = datetime.strptime(app_date_str.split()[0], '%Y-%m-%d')
                    if date_from:
                        from_date = datetime.strptime(date_from, '%Y-%m-%d')
                        if app_date < from_date:
                            continue
                    if date_to:
                        to_date = datetime.strptime(date_to, '%Y-%m-%d')
                        if app_date > to_date:
                            continue
                    filtered_applications.append(app)
                except (ValueError, IndexError):
                    continue
            all_applications = filtered_applications
        
        # Создаем файл экспорта
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_{timestamp}.xlsx"
        filepath = os.path.join(excel_integration.data_dir, filename)
        
        # Создаем Excel файл
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Экспорт заявок"
        
        # Заголовки
        headers = ["ОКЭ", "Время подачи заявки", "ФИО", "Табельный номер", "Должность", "Направление", "Информация о рейсе"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Данные
        for row, app in enumerate(all_applications, 2):
            ws.cell(row=row, column=1, value=app.get('oke', ''))
            ws.cell(row=row, column=2, value=app.get('Время подачи заявки', ''))
            ws.cell(row=row, column=3, value=app.get('ФИО', ''))
            ws.cell(row=row, column=4, value=app.get('Табельный номер', ''))
            ws.cell(row=row, column=5, value=app.get('Должность', ''))
            ws.cell(row=row, column=6, value=app.get('Направление', ''))
            ws.cell(row=row, column=7, value=app.get('Информация о рейсе', ''))
        
        wb.save(filepath)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'count': len(all_applications),
            'download_url': f'/api/download/{filename}'
        })
        
    except Exception as e:
        logger.error(f"Ошибка экспорта заявок: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download/<filename>')
def api_download_file(filename):
    """API для скачивания файлов"""
    try:
        from flask import send_file
        filepath = os.path.join(excel_integration.data_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': 'Файл не найден'}), 404
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Ошибка скачивания файла {filename}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Middleware для проверки авторизации
def require_auth(f):
    """Декоратор для проверки авторизации"""
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            token = request.cookies.get('auth_token')
        
        user = user_manager.get_user_by_token(token)
        if not user:
            return jsonify({'success': False, 'error': 'Не авторизован'}), 401
        
        request.current_user = user
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

def require_permission(permission: str):
    """Декоратор для проверки прав доступа"""
    def decorator(f):
        def decorated_function(*args, **kwargs):
            if not hasattr(request, 'current_user'):
                return jsonify({'success': False, 'error': 'Не авторизован'}), 401
            
            if not user_manager.has_permission(request.current_user, permission):
                return jsonify({'success': False, 'error': 'Недостаточно прав'}), 403
            
            return f(*args, **kwargs)
        decorated_function.__name__ = f.__name__
        return decorated_function
    return decorator

# API для авторизации
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    """API для входа в систему"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return jsonify({'success': False, 'error': 'Логин и пароль обязательны'}), 400
        
        # Получаем IP и User-Agent
        ip_address = request.remote_addr
        user_agent = request.headers.get('User-Agent', '')
        
        session = user_manager.authenticate_user(username, password, ip_address, user_agent)
        if not session:
            return jsonify({'success': False, 'error': 'Неверный логин или пароль'}), 401
        
        user = user_manager.users[session.user_id]
        
        response = jsonify({
            'success': True,
            'token': session.token,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role.value,
                'permissions': user_manager.get_user_permissions(user)
            }
        })
        
        # Устанавливаем cookie
        response.set_cookie('auth_token', session.token, max_age=24*60*60, httponly=True)
        
        return response
        
    except Exception as e:
        logger.error(f"Ошибка входа: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/auth/logout', methods=['POST'])
@require_auth
def api_logout():
    """API для выхода из системы"""
    try:
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            token = request.cookies.get('auth_token')
        
        success = user_manager.logout_user(token)
        
        response = jsonify({'success': success})
        response.set_cookie('auth_token', '', expires=0)
        
        return response
        
    except Exception as e:
        logger.error(f"Ошибка выхода: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/auth/me', methods=['GET'])
@require_auth
def api_get_current_user():
    """API для получения текущего пользователя"""
    try:
        user = request.current_user
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role.value,
                'status': user.status.value,
                'permissions': user_manager.get_user_permissions(user),
                'created_at': user.created_at.isoformat(),
                'last_login': user.last_login.isoformat() if user.last_login else None
            }
        })
    except Exception as e:
        logger.error(f"Ошибка получения пользователя: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# API для управления пользователями (только для админов)
@app.route('/api/users', methods=['GET'])
@require_auth
@require_permission('users.read')
def api_get_users():
    """API для получения списка пользователей"""
    try:
        users = user_manager.get_all_users()
        users_data = []
        
        for user in users:
            users_data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role.value,
                'status': user.status.value,
                'created_at': user.created_at.isoformat(),
                'last_login': user.last_login.isoformat() if user.last_login else None
            })
        
        return jsonify({
            'success': True,
            'users': users_data,
            'count': len(users_data)
        })
        
    except Exception as e:
        logger.error(f"Ошибка получения пользователей: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users', methods=['POST'])
@require_auth
@require_permission('users.create')
def api_create_user():
    """API для создания пользователя"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        full_name = data.get('full_name', '').strip()
        password = data.get('password', '').strip()
        role_str = data.get('role', 'user').strip()
        
        if not all([username, email, full_name, password]):
            return jsonify({'success': False, 'error': 'Все поля обязательны'}), 400
        
        try:
            role = UserRole(role_str)
        except ValueError:
            return jsonify({'success': False, 'error': 'Неверная роль'}), 400
        
        user = user_manager.create_user(username, email, full_name, password, role)
        if not user:
            return jsonify({'success': False, 'error': 'Пользователь с таким логином или email уже существует'}), 400
        
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role.value,
                'status': user.status.value,
                'created_at': user.created_at.isoformat()
            }
        })
        
    except Exception as e:
        logger.error(f"Ошибка создания пользователя: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/<user_id>', methods=['PUT'])
@require_auth
@require_permission('users.update')
def api_update_user(user_id):
    """API для обновления пользователя"""
    try:
        data = request.get_json()
        
        # Нельзя изменить роль администратора
        if user_id in user_manager.users and user_manager.users[user_id].role == UserRole.ADMIN:
            if 'role' in data and data['role'] != 'admin':
                return jsonify({'success': False, 'error': 'Нельзя изменить роль администратора'}), 400
        
        # Обновляем пользователя
        update_data = {}
        if 'full_name' in data:
            update_data['full_name'] = data['full_name'].strip()
        if 'email' in data:
            update_data['email'] = data['email'].strip()
        if 'role' in data:
            try:
                update_data['role'] = UserRole(data['role'])
            except ValueError:
                return jsonify({'success': False, 'error': 'Неверная роль'}), 400
        if 'status' in data:
            try:
                update_data['status'] = UserStatus(data['status'])
            except ValueError:
                return jsonify({'success': False, 'error': 'Неверный статус'}), 400
        
        success = user_manager.update_user(user_id, **update_data)
        if not success:
            return jsonify({'success': False, 'error': 'Пользователь не найден'}), 404
        
        user = user_manager.users[user_id]
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role.value,
                'status': user.status.value,
                'created_at': user.created_at.isoformat(),
                'last_login': user.last_login.isoformat() if user.last_login else None
            }
        })
        
    except Exception as e:
        logger.error(f"Ошибка обновления пользователя: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/<user_id>', methods=['DELETE'])
@require_auth
@require_permission('users.delete')
def api_delete_user(user_id):
    """API для удаления пользователя"""
    try:
        success = user_manager.delete_user(user_id)
        if not success:
            return jsonify({'success': False, 'error': 'Пользователь не найден или нельзя удалить'}), 404
        
        return jsonify({'success': True})
        
    except Exception as e:
        logger.error(f"Ошибка удаления пользователя: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API УВЕДОМЛЕНИЙ ====================

@app.route('/api/notifications', methods=['GET'])
@require_auth
def api_get_notifications():
    """API для получения уведомлений пользователя"""
    try:
        user_id = request.current_user.id
        unread_only = request.args.get('unread_only', 'true').lower() == 'true'
        
        notifications = notification_manager.get_user_notifications(user_id, unread_only)
        
        # Преобразуем в JSON-совместимый формат
        notifications_data = []
        for notif in notifications:
            notifications_data.append({
                'id': notif.id,
                'type': notif.type.value,
                'title': notif.title,
                'message': notif.message,
                'data': notif.data,
                'status': notif.status.value,
                'created_at': notif.created_at.isoformat(),
                'sent_at': notif.sent_at.isoformat() if notif.sent_at else None,
                'expires_at': notif.expires_at.isoformat() if notif.expires_at else None
            })
        
        return jsonify({
            'success': True,
            'notifications': notifications_data,
            'count': len(notifications_data)
        })
    except Exception as e:
        logger.error(f"Ошибка получения уведомлений: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/<notification_id>/read', methods=['POST'])
@require_auth
def api_mark_notification_read(notification_id):
    """API для отметки уведомления как прочитанного"""
    try:
        user_id = request.current_user.id
        success = notification_manager.mark_as_read(notification_id, user_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Уведомление отмечено как прочитанное'})
        else:
            return jsonify({'success': False, 'error': 'Уведомление не найдено'}), 404
    except Exception as e:
        logger.error(f"Ошибка отметки уведомления как прочитанного: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/subscribe', methods=['POST'])
@require_auth
def api_subscribe_notifications():
    """API для подписки на push-уведомления"""
    try:
        data = request.get_json()
        subscription_id = data.get('subscription_id')
        
        if not subscription_id:
            return jsonify({'success': False, 'error': 'ID подписки обязателен'}), 400
        
        user_id = request.current_user.id
        success = notification_manager.subscribe_user(user_id, subscription_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Подписка на уведомления активирована'})
        else:
            return jsonify({'success': False, 'error': 'Ошибка подписки'}), 400
    except Exception as e:
        logger.error(f"Ошибка подписки на уведомления: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/unsubscribe', methods=['POST'])
@require_auth
def api_unsubscribe_notifications():
    """API для отписки от push-уведомлений"""
    try:
        data = request.get_json()
        subscription_id = data.get('subscription_id')
        
        if not subscription_id:
            return jsonify({'success': False, 'error': 'ID подписки обязателен'}), 400
        
        user_id = request.current_user.id
        success = notification_manager.unsubscribe_user(user_id, subscription_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Подписка на уведомления отменена'})
        else:
            return jsonify({'success': False, 'error': 'Ошибка отписки'}), 400
    except Exception as e:
        logger.error(f"Ошибка отписки от уведомлений: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/stats', methods=['GET'])
@require_auth
@require_permission('reports.read')
def api_get_notification_stats():
    """API для получения статистики уведомлений"""
    try:
        stats = notification_manager.get_notification_stats()
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        logger.error(f"Ошибка получения статистики уведомлений: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/periods', methods=['GET'])
@require_auth
def api_get_notification_periods():
    """API для получения периодов приема заявок"""
    try:
        periods = notification_manager.get_active_application_periods()
        
        periods_data = []
        for period in periods:
            periods_data.append({
                'id': period.id,
                'name': period.name,
                'start_date': period.start_date.isoformat(),
                'end_date': period.end_date.isoformat(),
                'is_active': period.is_active,
                'reminder_hours': period.reminder_hours
            })
        
        return jsonify({
            'success': True,
            'periods': periods_data
        })
    except Exception as e:
        logger.error(f"Ошибка получения периодов приема заявок: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/application-periods', methods=['GET'])
@require_auth
def api_get_application_periods():
    """API для получения всех периодов подачи заявок"""
    try:
        periods = notification_manager.get_application_periods()
        return jsonify({
            'success': True,
            'periods': [period.to_dict() for period in periods]
        })
    except Exception as e:
        logger.error(f"Ошибка получения периодов заявок: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/public/application-periods', methods=['GET'])
def api_get_public_application_periods():
    """Публичный API для получения периодов подачи заявок (для ботов)"""
    try:
        periods = notification_manager.get_application_periods()
        # Возвращаем только базовую информацию без чувствительных данных
        public_periods = []
        for period in periods:
            # period это dataclass объект, работаем с атрибутами напрямую
            public_periods.append({
                'name': getattr(period, 'name', 'Период'),
                'start_date': getattr(period, 'start_date', '').isoformat() if hasattr(period, 'start_date') and getattr(period, 'start_date') else '',
                'end_date': getattr(period, 'end_date', '').isoformat() if hasattr(period, 'end_date') and getattr(period, 'end_date') else '',
                'is_active': getattr(period, 'is_active', False),
                'description': getattr(period, 'description', '') if hasattr(period, 'description') else ''
            })
        
        return jsonify({
            'success': True,
            'periods': public_periods,
            'total': len(public_periods)
        })
    except Exception as e:
        logger.error(f"Ошибка получения публичных периодов заявок: {e}")
        return jsonify({
            'success': False,
            'error': 'Не удалось получить информацию о периодах'
        }), 500

@app.route('/api/notifications/create', methods=['POST'])
@require_auth
@require_permission('applications.create')
def api_create_notification():
    """API для создания уведомления"""
    try:
        data = request.get_json()
        
        notification_type = NotificationType(data.get('type'))
        title = data.get('title')
        message = data.get('message')
        user_id = data.get('user_id')  # Если None - для всех пользователей
        expires_in_hours = data.get('expires_in_hours', 24)
        
        if not title or not message:
            return jsonify({'success': False, 'error': 'Заголовок и сообщение обязательны'}), 400
        
        notif_id = notification_manager.create_notification(
            type=notification_type,
            title=title,
            message=message,
            user_id=user_id,
            expires_in_hours=expires_in_hours
        )
        
        return jsonify({
            'success': True,
            'notification_id': notif_id,
            'message': 'Уведомление создано'
        })
    except Exception as e:
        logger.error(f"Ошибка создания уведомления: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== КОНСТРУКТОР УВЕДОМЛЕНИЙ ====================

@app.route('/api/notifications/schedules', methods=['GET'])
@require_auth
@require_permission('reports.read')
def api_get_notification_schedules():
    """API для получения расписания уведомлений"""
    try:
        schedules = notification_manager.get_notification_schedules()
        
        schedules_data = []
        for schedule in schedules:
            schedules_data.append({
                'id': schedule.id,
                'name': schedule.name,
                'type': schedule.type.value,
                'title': schedule.title,
                'message': schedule.message,
                'schedule_type': schedule.schedule_type.value,
                'schedule_config': schedule.schedule_config,
                'is_active': schedule.is_active,
                'created_at': schedule.created_at.isoformat(),
                'last_sent': schedule.last_sent.isoformat() if schedule.last_sent else None,
                'next_send': schedule.next_send.isoformat() if schedule.next_send else None
            })
        
        return jsonify({
            'success': True,
            'schedules': schedules_data
        })
    except Exception as e:
        logger.error(f"Ошибка получения расписания уведомлений: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/schedules', methods=['POST'])
@require_auth
@require_permission('applications.create')
def api_create_notification_schedule():
    """API для создания расписания уведомлений"""
    try:
        data = request.get_json()
        
        name = data.get('name')
        notification_type = NotificationType(data.get('type'))
        title = data.get('title')
        message = data.get('message')
        schedule_type = data.get('schedule_type')  # 'once', 'daily', 'weekly', 'monthly'
        schedule_config = data.get('schedule_config', {})
        is_active = data.get('is_active', True)
        
        if not all([name, title, message, schedule_type]):
            return jsonify({'success': False, 'error': 'Все поля обязательны'}), 400
        
        schedule_id = notification_manager.create_notification_schedule(
            name=name,
            notification_type=notification_type,
            title=title,
            message=message,
            schedule_type=schedule_type,
            schedule_config=schedule_config,
            is_active=is_active
        )
        
        return jsonify({
            'success': True,
            'schedule_id': schedule_id,
            'message': 'Расписание уведомлений создано'
        })
    except Exception as e:
        logger.error(f"Ошибка создания расписания уведомлений: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/schedules/<schedule_id>', methods=['PUT'])
@require_auth
@require_permission('applications.update')
def api_update_notification_schedule(schedule_id):
    """API для обновления расписания уведомлений"""
    try:
        data = request.get_json()
        
        success = notification_manager.update_notification_schedule(
            schedule_id=schedule_id,
            **data
        )
        
        if success:
            return jsonify({'success': True, 'message': 'Расписание обновлено'})
        else:
            return jsonify({'success': False, 'error': 'Расписание не найдено'}), 404
    except Exception as e:
        logger.error(f"Ошибка обновления расписания уведомлений: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/schedules/<schedule_id>', methods=['DELETE'])
@require_auth
@require_permission('applications.delete')
def api_delete_notification_schedule(schedule_id):
    """API для удаления расписания уведомлений"""
    try:
        success = notification_manager.delete_notification_schedule(schedule_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Расписание удалено'})
        else:
            return jsonify({'success': False, 'error': 'Расписание не найдено'}), 404
    except Exception as e:
        logger.error(f"Ошибка удаления расписания уведомлений: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notifications/schedules/<schedule_id>/test', methods=['POST'])
@require_auth
@require_permission('applications.create')
def api_test_notification_schedule(schedule_id):
    """API для тестирования расписания уведомлений"""
    try:
        success = notification_manager.test_notification_schedule(schedule_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Тестовое уведомление отправлено'})
        else:
            return jsonify({'success': False, 'error': 'Расписание не найдено'}), 404
    except Exception as e:
        logger.error(f"Ошибка тестирования расписания уведомлений: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# API для управления периодами подачи заявок (расширенные функции)
@app.route('/api/application-periods', methods=['POST'])
@require_auth
@require_permission('applications.create')
def api_create_application_period():
    """API для создания периода подачи заявок"""
    try:
        data = request.get_json()
        
        period = notification_manager.create_application_period(
            name=data['name'],
            start_date=data['start_date'],
            end_date=data['end_date'],
            description=data.get('description', ''),
            is_active=data.get('is_active', True)
        )
        
        if period:
            logger.info(f"✅ Период заявок создан: {period.name}")
            return jsonify({
                'success': True,
                'period': period.to_dict()
            })
        else:
            return jsonify({'success': False, 'error': 'Не удалось создать период'}), 400
            
    except Exception as e:
        logger.error(f"Ошибка создания периода заявок: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/application-periods/<period_id>', methods=['PUT'])
@require_auth
@require_permission('applications.create')
def api_update_application_period(period_id):
    """API для обновления периода подачи заявок"""
    try:
        data = request.get_json()
        
        success = notification_manager.update_application_period(
            period_id,
            name=data.get('name'),
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            description=data.get('description'),
            is_active=data.get('is_active')
        )
        
        if success:
            logger.info(f"✅ Период заявок обновлен: {period_id}")
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Период не найден'}), 404
            
    except Exception as e:
        logger.error(f"Ошибка обновления периода заявок: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/application-periods/<period_id>', methods=['DELETE'])
@require_auth
@require_permission('applications.create')
def api_delete_application_period(period_id):
    """API для удаления периода подачи заявок"""
    try:
        success = notification_manager.delete_application_period(period_id)
        
        if success:
            logger.info(f"✅ Период заявок удален: {period_id}")
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Период не найден'}), 404
            
    except Exception as e:
        logger.error(f"Ошибка удаления периода заявок: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Express мессенджер webhook endpoints
@app.route('/webhook/express', methods=['POST'])
def express_webhook():
    """Webhook для получения сообщений от Express мессенджера"""
    try:
        # Получаем данные
        data = request.get_json()
        signature = request.headers.get('X-Express-Signature', '')
        
        # Проверяем подпись
        bot = get_express_bot()
        if not bot.webhook_handler.verify_signature(request.data.decode('utf-8'), signature):
            logger.warning("Неверная подпись webhook от Express")
            return jsonify({'error': 'Unauthorized'}), 401
        
        # Парсим сообщение
        message = bot.webhook_handler.parse_message(data)
        if not message:
            return jsonify({'error': 'Invalid message format'}), 400
        
        # Обрабатываем сообщение
        response = bot.webhook_handler.handle_message(message)
        
        logger.info(f"Обработано сообщение от {message.user_id}: {message.text[:50]}...")
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"Ошибка обработки webhook: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/express/subscribe', methods=['POST'])
@require_auth
def api_express_subscribe():
    """API для подписки на уведомления Express"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        chat_id = data.get('chat_id')
        preferences = data.get('preferences', {})
        
        if not user_id or not chat_id:
            return jsonify({'success': False, 'error': 'user_id и chat_id обязательны'}), 400
        
        notification_service = get_notification_service()
        notification_service.subscribe_user(user_id, chat_id, preferences)
        
        logger.info(f"Пользователь {user_id} подписан на Express уведомления")
        return jsonify({'success': True, 'message': 'Подписка оформлена'})
        
    except Exception as e:
        logger.error(f"Ошибка подписки на Express: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/express/unsubscribe', methods=['POST'])
@require_auth
def api_express_unsubscribe():
    """API для отписки от уведомлений Express"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({'success': False, 'error': 'user_id обязателен'}), 400
        
        notification_service = get_notification_service()
        notification_service.unsubscribe_user(user_id)
        
        logger.info(f"Пользователь {user_id} отписан от Express уведомлений")
        return jsonify({'success': True, 'message': 'Отписка оформлена'})
        
    except Exception as e:
        logger.error(f"Ошибка отписки от Express: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/express/send-notification', methods=['POST'])
@require_auth
@require_permission('applications.create')
def api_express_send_notification():
    """API для отправки уведомления через Express"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        title = data.get('title')
        message = data.get('message')
        
        if not all([user_id, title, message]):
            return jsonify({'success': False, 'error': 'user_id, title и message обязательны'}), 400
        
        bot = get_express_bot()
        success = bot.send_notification(user_id, title, message)
        
        if success:
            logger.info(f"Уведомление отправлено пользователю {user_id}")
            return jsonify({'success': True, 'message': 'Уведомление отправлено'})
        else:
            return jsonify({'success': False, 'error': 'Не удалось отправить уведомление'}), 500
        
    except Exception as e:
        logger.error(f"Ошибка отправки уведомления Express: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/express/broadcast', methods=['POST'])
@require_auth
@require_permission('applications.create')
def api_express_broadcast():
    """API для массовой рассылки через Express"""
    try:
        data = request.get_json()
        title = data.get('title')
        message = data.get('message')
        user_filter = data.get('user_filter', {})
        
        if not all([title, message]):
            return jsonify({'success': False, 'error': 'title и message обязательны'}), 400
        
        notification_service = get_notification_service()
        
        # Простая фильтрация по ролям
        def filter_func(user_id, subscription):
            if 'roles' in user_filter:
                # Здесь можно добавить логику фильтрации по ролям
                pass
            return True
        
        result = notification_service.broadcast_notification(title, message, filter_func)
        
        logger.info(f"Рассылка завершена: {result['sent']} отправлено, {result['failed']} ошибок")
        return jsonify({
            'success': True, 
            'message': f"Рассылка завершена: {result['sent']} отправлено, {result['failed']} ошибок",
            'result': result
        })
        
    except Exception as e:
        logger.error(f"Ошибка массовой рассылки Express: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5002, debug=config.debug)
