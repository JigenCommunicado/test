#!/usr/bin/env python3
"""
Исправление Excel файла с тестовыми данными
"""

import sys
import os
sys.path.append('backend')

def fix_excel():
    print("🔧 Исправление Excel файла...")
    
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
        
        # Проверяем текущий файл
        excel_file = "data/все_заявки.xlsx"
        print(f"📁 Файл: {excel_file}")
        
        if os.path.exists(excel_file):
            try:
                wb = load_workbook(excel_file)
                print(f"📊 Текущие листы: {wb.sheetnames}")
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    print(f"  {sheet}: {ws.max_row} строк")
            except Exception as e:
                print(f"⚠️  Ошибка чтения: {e}")
        
        # Создаем новый файл с тестовыми данными
        print("\n📝 Создание тестовых данных...")
        wb = Workbook()
        wb.remove(wb.active)
        
        # Создаем лист ОКЭ 1
        ws = wb.create_sheet('ОКЭ 1')
        
        # Заголовки
        headers = [
            'Время подачи заявки', 
            'ФИО', 
            'Табельный номер', 
            'Должность', 
            'Направление', 
            'Информация о рейсе'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Тестовые данные
        test_data = [
            ['2024-09-09 10:30:00', 'Иванов И.И.', '12345', 'Инженер', 'Москва', '2024-09-15 08:00 SU-1234'],
            ['2024-09-09 11:15:00', 'Петров П.П.', '12346', 'Менеджер', 'СПб', '2024-09-16 14:30 SU-5678'],
            ['2024-09-09 12:00:00', 'Сидоров С.С.', '12347', 'Аналитик', 'Красноярск', '2024-09-17 16:45 SU-9012']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Сохраняем
        wb.save(excel_file)
        print(f"✅ Файл сохранен: {excel_file}")
        print(f"📊 Лист: ОКЭ 1")
        print(f"📋 Заявок: {len(test_data)}")
        
        # Проверяем через ExcelIntegration
        print("\n🔍 Проверка через ExcelIntegration...")
        from excel_integration import ExcelIntegration
        excel = ExcelIntegration(data_dir="data")
        
        apps = excel.read_applications('ОКЭ 1')
        print(f"📖 Прочитано заявок: {len(apps)}")
        
        if apps:
            print("📋 Заявки:")
            for i, app in enumerate(apps):
                print(f"  {i+1}. {app.get('ФИО')} - {app.get('Направление')}")
        
        return True
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        return False

if __name__ == "__main__":
    fix_excel()


