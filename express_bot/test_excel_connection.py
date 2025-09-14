#!/usr/bin/env python3
"""
Тест связи между сайтом и Excel файлом
"""

import sys
import os
sys.path.append('backend')

def test_excel_connection():
    print("🔍 Тестирование связи сайта с Excel файлом...")
    
    # 1. Проверяем файл
    excel_file = "data/все_заявки.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Файл {excel_file} не найден!")
        return False
    
    print(f"✅ Файл {excel_file} существует")
    
    # 2. Проверяем через openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        print(f"📊 Листы: {wb.sheetnames}")
        
        # Проверяем каждый лист
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📋 {sheet_name}: {ws.max_row} строк")
            
            if ws.max_row > 1:
                # Показываем заголовки
                headers = []
                for col in range(1, ws.max_column + 1):
                    headers.append(ws.cell(1, col).value)
                print(f"    Заголовки: {headers}")
                
                # Показываем данные
                if ws.max_row > 2:
                    print(f"    Есть данные: {ws.max_row - 1} строк")
                    # Показываем первую строку данных
                    first_data = []
                    for col in range(1, ws.max_column + 1):
                        first_data.append(ws.cell(2, col).value)
                    print(f"    Первая строка: {first_data}")
                else:
                    print(f"    Нет данных (только заголовки)")
            else:
                print(f"    Пустой лист")
        
    except Exception as e:
        print(f"❌ Ошибка чтения Excel: {e}")
        return False
    
    # 3. Проверяем через ExcelIntegration
    try:
        from excel_integration import ExcelIntegration
        excel = ExcelIntegration(data_dir="data")
        
        print(f"\n🔧 Тест ExcelIntegration:")
        print(f"  Файл загружен: {excel.workbook is not None}")
        
        if excel.workbook:
            print(f"  Листы в ExcelIntegration: {excel.workbook.sheetnames}")
        
        # Тестируем чтение заявок
        oke_list = ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']
        total_apps = 0
        
        for oke in oke_list:
            apps = excel.read_applications(oke)
            print(f"  {oke}: {len(apps)} заявок")
            total_apps += len(apps)
        
        print(f"\n📊 Итого заявок: {total_apps}")
        
        if total_apps == 0:
            print("⚠️  Проблема: заявки не найдены!")
            print("   Возможные причины:")
            print("   1. Листы в Excel не соответствуют названиям ОКЭ")
            print("   2. В листах нет данных (только заголовки)")
            print("   3. Формат данных не соответствует ожидаемому")
        else:
            print("✅ Заявки найдены!")
        
        return total_apps > 0
        
    except Exception as e:
        print(f"❌ Ошибка ExcelIntegration: {e}")
        return False

if __name__ == "__main__":
    test_excel_connection()


