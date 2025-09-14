#!/usr/bin/env python3
"""
Скрипт для отладки Excel файла
"""

import sys
import os
sys.path.append('backend')

def check_excel_file():
    print("🔍 Проверка Excel файла...")
    
    # Проверяем существование файла
    excel_file = "data/все_заявки.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Файл {excel_file} не найден!")
        return
    
    print(f"✅ Файл {excel_file} существует")
    
    # Проверяем с помощью openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        print(f"📊 Листы в файле: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  📋 {sheet_name}: {ws.max_row} строк, {ws.max_column} колонок")
            
            # Показываем первые несколько строк
            if ws.max_row > 1:
                print(f"    Заголовки: {[ws.cell(1, col).value for col in range(1, ws.max_column + 1)]}")
                if ws.max_row > 2:
                    print(f"    Первая строка данных: {[ws.cell(2, col).value for col in range(1, ws.max_column + 1)]}")
        
    except Exception as e:
        print(f"❌ Ошибка чтения Excel файла: {e}")
        return
    
    # Проверяем через ExcelIntegration
    try:
        from excel_integration import ExcelIntegration
        excel = ExcelIntegration(data_dir="data")
        
        print(f"\n🔧 Проверка через ExcelIntegration:")
        print(f"  Файл загружен: {excel.workbook is not None}")
        print(f"  Путь к файлу: {excel.master_file_path}")
        
        # Проверяем каждый ОКЭ
        oke_list = ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']
        total_apps = 0
        
        for oke in oke_list:
            apps = excel.read_applications(oke)
            print(f"  {oke}: {len(apps)} заявок")
            total_apps += len(apps)
            
            if apps:
                print(f"    Пример: {apps[0]}")
        
        print(f"\n📊 Всего заявок: {total_apps}")
        
    except Exception as e:
        print(f"❌ Ошибка ExcelIntegration: {e}")

if __name__ == "__main__":
    check_excel_file()


