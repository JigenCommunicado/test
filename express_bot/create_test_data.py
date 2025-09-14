#!/usr/bin/env python3
"""
Создание тестовых данных в Excel файле
"""

import sys
import os
from datetime import datetime
sys.path.append('backend')

def create_test_data():
    print("📝 Создание тестовых данных...")
    
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Заголовки
        headers = [
            "Время подачи заявки",
            "ФИО", 
            "Табельный номер",
            "Должность",
            "Направление",
            "Информация о рейсе"
        ]
        
        # Тестовые данные
        test_data = [
            [
                "2024-09-09 10:30:00",
                "Иванов Иван Иванович",
                "12345",
                "Инженер",
                "Москва",
                "2024-09-15 08:00 SU-1234"
            ],
            [
                "2024-09-09 11:15:00",
                "Петров Петр Петрович",
                "12346",
                "Менеджер",
                "Санкт-Петербург",
                "2024-09-16 14:30 SU-5678"
            ],
            [
                "2024-09-09 12:00:00",
                "Сидоров Сидор Сидорович",
                "12347",
                "Аналитик",
                "Красноярск",
                "2024-09-17 16:45 SU-9012"
            ]
        ]
        
        # Создаем или загружаем файл
        excel_file = "data/все_заявки.xlsx"
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            print("📂 Загружен существующий файл")
        else:
            wb = Workbook()
            print("📄 Создан новый файл")
        
        # Создаем лист для ОКЭ 1
        oke_name = "ОКЭ 1"
        if oke_name in wb.sheetnames:
            ws = wb[oke_name]
            # Очищаем существующие данные
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(oke_name)
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Записываем тестовые данные
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(excel_file)
        print(f"✅ Тестовые данные сохранены в {excel_file}")
        print(f"   Лист: {oke_name}")
        print(f"   Заявок: {len(test_data)}")
        
        return True
        
    except Exception as e:
        print(f"❌ Ошибка создания тестовых данных: {e}")
        return False

if __name__ == "__main__":
    create_test_data()


