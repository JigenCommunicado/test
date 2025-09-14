#!/root/test/express_bot/venv/bin/python3
"""
Простой Flask сервер для админ панели
"""

from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
import os
import sys
import time
import threading
import asyncio
import random
import psutil

# Добавляем путь к модулям бота
sys.path.append('/root/test/express_bot')
sys.path.append('/root/test/express_bot/backend')
from config_manager import config_manager

# Импортируем ExcelIntegration для работы с данными
try:
    from excel_integration import ExcelIntegration
except ImportError:
    ExcelIntegration = None

app = Flask(__name__)
CORS(app)

# Простая конфигурация
config = {
    'bot_id': '00c46d64-1127-5a96-812d-3d8b27c58b99',
    'bot_name': 'Flight Booking Bot',
    'bot_description': 'Бот для подачи заявок на командировочные рейсы',
    'host': '0.0.0.0',
    'port': 5006
}

# Загружаем конфигурацию из файла
try:
    import json
    # Убрали импорт socket - больше не нужен
    with open('config.json', 'r', encoding='utf-8') as f:
        config_data = json.load(f)
        if 'ports' in config_data:
            config['port'] = config_data['ports'].get('admin_server', 8080)
        if 'bot_settings' in config_data:
            config['bot_id'] = config_data['bot_settings'].get('bot_id', config['bot_id'])
            config['bot_name'] = config_data['bot_settings'].get('bot_name', config['bot_name'])
            config['bot_description'] = config_data['bot_settings'].get('bot_description', config['bot_description'])
except Exception as e:
    print(f"⚠️ Не удалось загрузить config.json: {e}")
    print("Используем значения по умолчанию")

# Используем фиксированный порт из конфигурации

@app.route('/')
def admin_panel():
    """Главная страница - админ панель"""
    try:
        # Читаем файл админ панели
        admin_panel_path = '/root/test/express_bot/frontend/admin_panel.html'
        with open(admin_panel_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Заменяем URL API на правильный
        content = content.replace(
            'const SMARTAPP_URL = window.location.origin;',
            f'const SMARTAPP_URL = \'http://localhost:{config["port"]}\';'
        )
        
        return content
        
    except Exception as e:
        return f"<h1>Ошибка загрузки админ панели</h1><p>Ошибка: {str(e)}</p>", 500

@app.route('/admin')
def admin_panel_alt():
    """Альтернативный путь к админ панели"""
    return admin_panel()

@app.route('/health')
def health():
    """Проверка здоровья сервера"""
    return jsonify({
        "status": "ok",
        "service": "Admin Panel Server",
        "bot_id": config['bot_id']
    })

@app.route('/manifest')
def manifest():
    """Манифест бота"""
    manifest = {
        "name": config['bot_name'],
        "version": "1.0.0",
        "description": config['bot_description'],
        "icon": "✈️",
        "color": "#0088cc",
        "author": "Express Bot Team",
        "webhook": "/webhook",
        "commands": [
            {
                "command": "/start",
                "description": "Начать работу с ботом"
            },
            {
                "command": "/new",
                "description": "Подать новую заявку на рейс"
            },
            {
                "command": "/my",
                "description": "Мои заявки"
            },
            {
                "command": "/help",
                "description": "Справка"
            },
            {
                "command": "/status",
                "description": "Статус системы"
            }
        ]
    }
    return jsonify(manifest)

# API обработчики для админ панели
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    """API: Авторизация"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        # Простая проверка (в реальном проекте нужна база данных)
        if username == 'admin' and password == 'admin':
            return jsonify({
                "success": True,
                "token": "demo_token_123",
                "user": {
                    "id": 1,
                    "username": "admin",
                    "email": "admin@example.com",
                    "full_name": "Администратор",
                    "role": "admin",
                    "status": "active",
                    "permissions": ["all"]
                }
            })
        else:
            return jsonify({
                "success": False,
                "error": "Неверные учетные данные"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": "Ошибка сервера"
        })

@app.route('/api/auth/me', methods=['GET'])
def api_me():
    """API: Информация о текущем пользователе"""
    return jsonify({
        "success": True,
        "user": {
            "id": 1,
            "username": "admin",
            "email": "admin@example.com",
            "full_name": "Администратор",
            "role": "admin",
            "status": "active",
            "permissions": ["all"]
        }
    })

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    """API: Выход из системы"""
    return jsonify({"success": True})

@app.route('/api/users', methods=['GET'])
def api_users():
    """API: Список пользователей"""
    users = [
        {
            "id": 1,
            "username": "admin",
            "email": "admin@example.com",
            "full_name": "Администратор",
            "role": "admin",
            "status": "active",
            "created_at": "2025-01-01T00:00:00Z"
        },
        {
            "id": 2,
            "username": "manager1",
            "email": "manager@example.com",
            "full_name": "Менеджер",
            "role": "manager",
            "status": "active",
            "created_at": "2025-01-02T00:00:00Z"
        }
    ]
    return jsonify({"success": True, "users": users})

@app.route('/api/users', methods=['POST'])
def api_create_user():
    """API: Создание пользователя"""
    try:
        data = request.get_json()
        # В реальном проекте здесь была бы запись в базу данных
        return jsonify({"success": True, "message": "Пользователь создан"})
    except Exception as e:
        return jsonify({"success": False, "error": "Ошибка создания пользователя"})

@app.route('/api/users/<user_id>', methods=['DELETE'])
def api_delete_user(user_id):
    """API: Удаление пользователя"""
    return jsonify({"success": True, "message": f"Пользователь {user_id} удален"})

@app.route('/api/notifications/schedules', methods=['GET'])
def api_schedules():
    """API: Список расписаний уведомлений"""
    schedules = [
        {
            "id": 1,
            "name": "Ежемесячное напоминание",
            "type": "application_reminder",
            "schedule_type": "monthly",
            "is_active": True,
            "next_send": "2025-10-01T09:00:00Z"
        }
    ]
    return jsonify({"success": True, "schedules": schedules})

@app.route('/api/notifications/schedules', methods=['POST'])
def api_create_schedule():
    """API: Создание расписания уведомлений"""
    try:
        data = request.get_json()
        return jsonify({"success": True, "message": "Расписание создано"})
    except Exception as e:
        return jsonify({"success": False, "error": "Ошибка создания расписания"})

# ==================== НОВЫЕ API ДЛЯ УПРАВЛЕНИЯ БОТОМ ====================

@app.route('/api/bot/status', methods=['GET'])
def api_bot_status():
    """API: Статус бота"""
    try:
        import subprocess
        import os
        from datetime import datetime
        
        # Проверяем, запущен ли процесс бота
        bot_processes = []
        for proc in psutil.process_iter(['pid', 'name', 'cmdline', 'create_time']):
            try:
                if proc.info['name'] == 'python3' and proc.info['cmdline']:
                    cmdline = ' '.join(proc.info['cmdline'])
                    if 'express_bot.py' in cmdline:
                        bot_processes.append({
                            'pid': proc.info['pid'],
                            'create_time': datetime.fromtimestamp(proc.info['create_time']).isoformat(),
                            'status': 'running'
                        })
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        
        # Получаем информацию о системе
        system_info = {
            'cpu_percent': psutil.cpu_percent(interval=1),
            'memory_percent': psutil.virtual_memory().percent,
            'disk_percent': psutil.disk_usage('/').percent,
            'uptime': datetime.now().isoformat()
        }
        
        return jsonify({
            "success": True,
            "bot_status": "running" if bot_processes else "stopped",
            "processes": bot_processes,
            "system": system_info
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения статуса: {str(e)}"
        })

@app.route('/api/bot/start', methods=['POST'])
def api_bot_start():
    """API: Запуск бота"""
    try:
        import subprocess
        import os
        
        # Проверяем, не запущен ли уже бот
        for proc in psutil.process_iter(['name', 'cmdline']):
            try:
                if proc.info['name'] == 'python3' and proc.info['cmdline']:
                    cmdline = ' '.join(proc.info['cmdline'])
                    if 'express_bot.py' in cmdline:
                        return jsonify({
                            "success": False,
                            "error": "Бот уже запущен"
                        })
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        
        # Запускаем бота в фоновом режиме
        bot_dir = '/root/test/express_bot'
        result = subprocess.Popen(
            ['python3', 'express_bot.py'],
            cwd=bot_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        
        return jsonify({
            "success": True,
            "message": "Бот запущен",
            "pid": result.pid
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка запуска бота: {str(e)}"
        })

@app.route('/api/bot/stop', methods=['POST'])
def api_bot_stop():
    """API: Остановка бота"""
    try:
        import subprocess
        
        # Находим и останавливаем процессы бота
        stopped_count = 0
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                if proc.info['name'] == 'python3' and proc.info['cmdline']:
                    cmdline = ' '.join(proc.info['cmdline'])
                    if 'express_bot.py' in cmdline:
                        proc.terminate()
                        stopped_count += 1
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        
        return jsonify({
            "success": True,
            "message": f"Остановлено процессов: {stopped_count}"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка остановки бота: {str(e)}"
        })

@app.route('/api/statistics', methods=['GET'])
def api_statistics():
    """API: Статистика системы"""
    try:
        # Здесь должна быть реальная статистика из базы данных
        # Пока возвращаем заглушку
        stats = {
            "applications": {
                "total": 156,
                "today": 12,
                "this_week": 89,
                "this_month": 156
            },
            "users": {
                "total": 45,
                "active_today": 23,
                "active_this_week": 38
            },
            "popular_directions": [
                {"name": "Москва", "count": 45},
                {"name": "Санкт-Петербург", "count": 32},
                {"name": "Красноярск", "count": 28},
                {"name": "Сочи", "count": 15}
            ],
            "system_health": {
                "bot_uptime": "2 дня 5 часов",
                "last_error": None,
                "response_time": "120ms"
            }
        }
        
        return jsonify({
            "success": True,
            "statistics": stats
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения статистики: {str(e)}"
        })

@app.route('/api/logs', methods=['GET'])
def api_logs():
    """API: Получение логов"""
    try:
        import os
        from datetime import datetime, timedelta
        
        # Параметры запроса
        lines = request.args.get('lines', 100, type=int)
        level = request.args.get('level', 'all')
        
        # Путь к файлу логов
        log_file = '/root/test/express_bot/bot.log'
        
        if not os.path.exists(log_file):
            return jsonify({
                "success": True,
                "logs": ["Файл логов не найден"],
                "total_lines": 0
            })
        
        # Читаем последние N строк
        with open(log_file, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()
        
        # Фильтруем по уровню
        if level != 'all':
            filtered_lines = [line for line in all_lines if level.upper() in line]
        else:
            filtered_lines = all_lines
        
        # Берем последние строки
        recent_lines = filtered_lines[-lines:] if len(filtered_lines) > lines else filtered_lines
        
        return jsonify({
            "success": True,
            "logs": [line.strip() for line in recent_lines],
            "total_lines": len(filtered_lines),
            "level": level
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения логов: {str(e)}"
        })

# ==================== API ДЛЯ УПРАВЛЕНИЯ КОНФИГУРАЦИЕЙ ====================

@app.route('/api/config', methods=['GET'])
def api_get_config():
    """API: Получение всей конфигурации"""
    try:
        config = config_manager.config
        validation = config_manager.validate_config()
        
        return jsonify({
            "success": True,
            "config": config,
            "validation": validation
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения конфигурации: {str(e)}"
        })

@app.route('/api/config', methods=['POST'])
def api_save_config():
    """API: Сохранение конфигурации"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                "success": False,
                "error": "Данные конфигурации не предоставлены"
            })
        
        # Обновляем конфигурацию
        config_manager.config = data
        
        # Валидируем
        validation = config_manager.validate_config()
        if validation["errors"]:
            return jsonify({
                "success": False,
                "error": "Ошибки валидации",
                "validation": validation
            })
        
        # Сохраняем
        if config_manager.save_config():
            return jsonify({
                "success": True,
                "message": "Конфигурация сохранена",
                "validation": validation
            })
        else:
            return jsonify({
                "success": False,
                "error": "Ошибка сохранения конфигурации"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка сохранения конфигурации: {str(e)}"
        })

@app.route('/api/config/directions', methods=['GET'])
def api_get_directions():
    """API: Получение направлений"""
    try:
        directions = config_manager.get("directions", {})
        return jsonify({
            "success": True,
            "directions": directions
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения направлений: {str(e)}"
        })

@app.route('/api/config/directions', methods=['POST'])
def api_add_direction():
    """API: Добавление направления"""
    try:
        data = request.get_json()
        location = data.get('location')
        direction = data.get('direction')
        
        if not location or not direction:
            return jsonify({
                "success": False,
                "error": "Локация и направление обязательны"
            })
        
        if config_manager.add_direction(location, direction):
            config_manager.save_config()
            return jsonify({
                "success": True,
                "message": f"Направление '{direction}' добавлено для {location}"
            })
        else:
            return jsonify({
                "success": False,
                "error": "Направление уже существует"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка добавления направления: {str(e)}"
        })

@app.route('/api/config/directions', methods=['DELETE'])
def api_remove_direction():
    """API: Удаление направления"""
    try:
        location = request.args.get('location')
        direction = request.args.get('direction')
        
        if not location or not direction:
            return jsonify({
                "success": False,
                "error": "Локация и направление обязательны"
            })
        
        if config_manager.remove_direction(location, direction):
            config_manager.save_config()
            return jsonify({
                "success": True,
                "message": f"Направление '{direction}' удалено для {location}"
            })
        else:
            return jsonify({
                "success": False,
                "error": "Направление не найдено"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка удаления направления: {str(e)}"
        })

@app.route('/api/config/oke', methods=['GET'])
def api_get_oke():
    """API: Получение ОКЭ"""
    try:
        oke = config_manager.get("oke_by_location", {})
        return jsonify({
            "success": True,
            "oke": oke
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения ОКЭ: {str(e)}"
        })

@app.route('/api/config/oke', methods=['POST'])
def api_add_oke():
    """API: Добавление ОКЭ"""
    try:
        data = request.get_json()
        location = data.get('location')
        oke = data.get('oke')
        
        if not location or not oke:
            return jsonify({
                "success": False,
                "error": "Локация и ОКЭ обязательны"
            })
        
        if config_manager.add_oke(location, oke):
            config_manager.save_config()
            return jsonify({
                "success": True,
                "message": f"ОКЭ '{oke}' добавлен для {location}"
            })
        else:
            return jsonify({
                "success": False,
                "error": "ОКЭ уже существует"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка добавления ОКЭ: {str(e)}"
        })

@app.route('/api/config/oke', methods=['DELETE'])
def api_remove_oke():
    """API: Удаление ОКЭ"""
    try:
        location = request.args.get('location')
        oke = request.args.get('oke')
        
        if not location or not oke:
            return jsonify({
                "success": False,
                "error": "Локация и ОКЭ обязательны"
            })
        
        if config_manager.remove_oke(location, oke):
            config_manager.save_config()
            return jsonify({
                "success": True,
                "message": f"ОКЭ '{oke}' удален для {location}"
            })
        else:
            return jsonify({
                "success": False,
                "error": "ОКЭ не найден"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка удаления ОКЭ: {str(e)}"
        })

@app.route('/api/config/positions', methods=['GET'])
def api_get_positions():
    """API: Получение должностей"""
    try:
        positions = config_manager.get_positions()
        return jsonify({
            "success": True,
            "positions": positions
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения должностей: {str(e)}"
        })

@app.route('/api/config/positions', methods=['POST'])
def api_add_position():
    """API: Добавление должности"""
    try:
        data = request.get_json()
        position = data.get('position')
        
        if not position:
            return jsonify({
                "success": False,
                "error": "Должность обязательна"
            })
        
        if config_manager.add_position(position):
            config_manager.save_config()
            return jsonify({
                "success": True,
                "message": f"Должность '{position}' добавлена"
            })
        else:
            return jsonify({
                "success": False,
                "error": "Должность уже существует"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка добавления должности: {str(e)}"
        })

@app.route('/api/config/positions', methods=['DELETE'])
def api_remove_position():
    """API: Удаление должности"""
    try:
        position = request.args.get('position')
        
        if not position:
            return jsonify({
                "success": False,
                "error": "Должность обязательна"
            })
        
        if config_manager.remove_position(position):
            config_manager.save_config()
            return jsonify({
                "success": True,
                "message": f"Должность '{position}' удалена"
            })
        else:
            return jsonify({
                "success": False,
                "error": "Должность не найдена"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка удаления должности: {str(e)}"
        })

@app.route('/api/config/backup', methods=['GET'])
def api_get_backups():
    """API: Получение списка резервных копий"""
    try:
        backups = config_manager.get_backup_files()
        return jsonify({
            "success": True,
            "backups": backups
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения резервных копий: {str(e)}"
        })

@app.route('/api/config/backup/restore', methods=['POST'])
def api_restore_backup():
    """API: Восстановление из резервной копии"""
    try:
        data = request.get_json()
        backup_file = data.get('backup_file')
        
        if not backup_file:
            return jsonify({
                "success": False,
                "error": "Файл резервной копии обязателен"
            })
        
        if config_manager.restore_backup(backup_file):
            return jsonify({
                "success": True,
                "message": "Конфигурация восстановлена из резервной копии"
            })
        else:
            return jsonify({
                "success": False,
                "error": "Ошибка восстановления из резервной копии"
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка восстановления: {str(e)}"
        })

# ==================== API ДЛЯ РАБОТЫ С ЗАЯВКАМИ ====================

@app.route('/api/applications', methods=['GET'])
def api_get_applications():
    """API: Получение списка заявок из реальных данных"""
    try:
        # Используем ExcelIntegration для работы с реальными данными
        if ExcelIntegration is None:
            return jsonify({
                "success": False,
                "error": "ExcelIntegration не доступен"
            })
        
        excel_integration = ExcelIntegration(data_dir="data")
        all_applications = []
        
        print(f"🔍 Ищем заявки в директории: {os.path.join(os.path.dirname(__file__), 'data')}")
        
        # Получаем заявки из всех ОКЭ
        for oke in ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']:
            try:
                print(f"📋 Читаем заявки из {oke}...")
                applications = excel_integration.read_applications(oke)
                print(f"📊 Найдено {len(applications)} заявок в {oke}")
                
                for i, app in enumerate(applications):
                    # Преобразуем данные в формат для админ панели
                    all_applications.append({
                        "id": f"{oke}_{i+1}",
                        "full_name": app.get('ФИО', ''),
                        "direction": app.get('Направление', ''),
                        "created_at": app.get('Время подачи заявки', ''),
                        "departure_date": app.get('Информация о рейсе', '').split()[0] if app.get('Информация о рейсе') else '',
                        "position": app.get('Должность', ''),
                        "tab_num": app.get('Табельный номер', ''),
                        "oke": oke
                    })
            except Exception as e:
                print(f"❌ Ошибка чтения заявок из {oke}: {e}")
                continue
        
        print(f"✅ Всего заявок найдено: {len(all_applications)}")
        
        # Сортируем по дате создания (новые сначала)
        all_applications.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        return jsonify({
            "success": True,
            "applications": all_applications,
            "total": len(all_applications)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения заявок: {str(e)}"
        })

@app.route('/api/applications/<int:app_id>/status', methods=['PUT'])
def api_update_application_status(app_id):
    """API: Обновление статуса заявки"""
    try:
        data = request.get_json()
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({
                "success": False,
                "error": "Статус не указан"
            })
        
        # Здесь должна быть логика обновления статуса в базе данных
        # Пока просто возвращаем успех
        
        return jsonify({
            "success": True,
            "message": f"Статус заявки #{app_id} обновлен на '{new_status}'"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка обновления статуса: {str(e)}"
        })

@app.route('/api/applications/export', methods=['POST'])
def api_export_applications():
    """API: Экспорт заявок в Excel"""
    try:
        data = request.get_json()
        filters = data.get('filters', {})
        
        # Создаем Excel файл
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import datetime
        
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Все заявки"
        
        # Настройка заголовков
        headers = ['ID', 'ФИО', 'Направление', 'Дата создания', 'Дата вылета']
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Получаем реальные данные из Excel
        if ExcelIntegration is None:
            return jsonify({
                "success": False,
                "error": "ExcelIntegration не доступен"
            })
        
        excel_integration = ExcelIntegration(data_dir="data")
        applications = []
        
        # Получаем заявки из всех ОКЭ
        for oke in ['ОКЭ 1', 'ОКЭ 2', 'ОКЭ 3', 'ОЛСиТ', 'ОКЭ 4', 'ОКЭ 5', 'ОКЭ Красноярск', 'ОКЭ Сочи']:
            try:
                oke_applications = excel_integration.read_applications(oke)
                for i, app in enumerate(oke_applications):
                    # Преобразуем дату в формат дд.мм.гггг
                    created_date = app.get('Время подачи заявки', '')
                    if created_date:
                        try:
                            # Парсим дату из формата "2024-01-15 10:30:00"
                            dt = datetime.datetime.strptime(created_date.split()[0], '%Y-%m-%d')
                            created_date = dt.strftime('%d.%m.%Y')
                        except:
                            created_date = created_date.split()[0] if ' ' in created_date else created_date
                    
                    departure_date = app.get('Информация о рейсе', '').split()[0] if app.get('Информация о рейсе') else ''
                    
                    applications.append({
                        'id': f"{oke}_{i+1}",
                        'full_name': app.get('ФИО', ''),
                        'direction': app.get('Направление', ''),
                        'created_at': created_date,
                        'departure_date': departure_date
                    })
            except Exception as e:
                print(f"Ошибка чтения заявок из {oke}: {e}")
                continue
        
        # Записываем данные
        for row, app in enumerate(applications, 2):
            ws.cell(row=row, column=1, value=app['id'])
            ws.cell(row=row, column=2, value=app['full_name'])
            ws.cell(row=row, column=3, value=app['direction'])
            ws.cell(row=row, column=4, value=app['created_at'])
            ws.cell(row=row, column=5, value=app['departure_date'])
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        filename = "все_заявки.xlsx"
        filepath = f"/root/test/express_bot/data/{filename}"
        
        # Создаем директорию если не существует
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Сохраняем файл
        wb.save(filepath)
        
        # Читаем файл для отправки
        with open(filepath, 'rb') as f:
            file_data = f.read()
        
        from flask import Response
        return Response(
            file_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка экспорта: {str(e)}"
        })

# ==================== API ДЛЯ СТРЕСС-ТЕСТИРОВАНИЯ ====================

# Глобальные переменные для стресс-тестирования
stress_test_running = False
stress_test_data = {
    'status': 'idle',
    'test_type': None,
    'total': 0,
    'completed': 0,
    'successful': 0,
    'errors': 0,
    'start_time': None,
    'end_time': None,
    'response_times': [],
    'concurrency': 0,
    'delay': 0
}

@app.route('/api/stress-test/start', methods=['POST'])
def api_start_stress_test():
    """API: Запуск стресс-теста"""
    global stress_test_running, stress_test_data
    
    if stress_test_running:
        return jsonify({
            "success": False,
            "error": "Стресс-тест уже выполняется"
        })
    
    try:
        data = request.get_json()
        test_type = data.get('test_type', 'applications')
        count = data.get('count', 100)
        delay = data.get('delay', 100)
        concurrency = data.get('concurrency', 5)
        
        # Инициализируем данные теста
        stress_test_data = {
            'status': 'running',
            'test_type': test_type,
            'total': count,
            'completed': 0,
            'successful': 0,
            'errors': 0,
            'start_time': time.time(),
            'end_time': None,
            'response_times': [],
            'concurrency': concurrency,
            'delay': delay
        }
        
        stress_test_running = True
        
        # Запускаем тест в отдельном потоке
        test_thread = threading.Thread(target=run_stress_test, args=(test_type, count, delay, concurrency))
        test_thread.daemon = True
        test_thread.start()
        
        return jsonify({
            "success": True,
            "message": f"Стресс-тест '{test_type}' запущен ({count} операций)"
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка запуска стресс-теста: {str(e)}"
        })

@app.route('/api/stress-test/stop', methods=['POST'])
def api_stop_stress_test():
    """API: Остановка стресс-теста"""
    global stress_test_running, stress_test_data
    
    if not stress_test_running:
        return jsonify({
            "success": False,
            "error": "Стресс-тест не выполняется"
        })
    
    stress_test_running = False
    stress_test_data['status'] = 'stopped'
    stress_test_data['end_time'] = time.time()
    
    return jsonify({
        "success": True,
        "message": "Стресс-тест остановлен"
    })

@app.route('/api/stress-test/status', methods=['GET'])
def api_get_stress_test_status():
    """API: Получение статуса стресс-теста"""
    global stress_test_data
    
    try:
        # Вычисляем метрики
        success_rate = 0
        avg_response_time = 0
        duration = 0
        rps = 0
        
        if stress_test_data['completed'] > 0:
            success_rate = (stress_test_data['successful'] / stress_test_data['completed']) * 100
        
        if stress_test_data['response_times']:
            avg_response_time = sum(stress_test_data['response_times']) / len(stress_test_data['response_times'])
        
        if stress_test_data['start_time']:
            end_time = stress_test_data['end_time'] or time.time()
            duration = end_time - stress_test_data['start_time']
            
            if duration > 0:
                rps = stress_test_data['completed'] / duration
        
        # Детальная статистика
        details = {
            'successful': stress_test_data['successful'],
            'errors': stress_test_data['errors'],
            'min_time': min(stress_test_data['response_times']) if stress_test_data['response_times'] else 0,
            'max_time': max(stress_test_data['response_times']) if stress_test_data['response_times'] else 0,
            'rps': round(rps, 2),
            'concurrency': stress_test_data['concurrency']
        }
        
        return jsonify({
            "success": True,
            "status": stress_test_data['status'],
            "test_type": stress_test_data['test_type'],
            "total": stress_test_data['total'],
            "completed": stress_test_data['completed'],
            "success_rate": round(success_rate, 2),
            "avg_response_time": round(avg_response_time, 2),
            "duration": round(duration, 2),
            "details": details
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Ошибка получения статуса: {str(e)}"
        })

def run_stress_test(test_type, count, delay, concurrency):
    """Выполнение стресс-теста"""
    global stress_test_running, stress_test_data
    
    async def simulate_operation(operation_id):
        """Симуляция одной операции"""
        start_time = time.time()
        
        try:
            # Симуляция задержки
            await asyncio.sleep(delay / 1000.0)
            
            # Симуляция различных типов операций
            if test_type == 'applications':
                # Симуляция создания заявки
                await simulate_create_application()
            elif test_type == 'users':
                # Симуляция создания пользователя
                await simulate_create_user()
            elif test_type == 'notifications':
                # Симуляция отправки уведомления
                await simulate_send_notification()
            elif test_type == 'mixed':
                # Смешанный тест
                operations = ['applications', 'users', 'notifications']
                operation = random.choice(operations)
                if operation == 'applications':
                    await simulate_create_application()
                elif operation == 'users':
                    await simulate_create_user()
                else:
                    await simulate_send_notification()
            
            # Успешное выполнение
            response_time = (time.time() - start_time) * 1000  # в миллисекундах
            stress_test_data['successful'] += 1
            stress_test_data['response_times'].append(response_time)
            
        except Exception as e:
            # Ошибка выполнения
            stress_test_data['errors'] += 1
            print(f"Ошибка операции {operation_id}: {e}")
        
        finally:
            stress_test_data['completed'] += 1
    
    async def simulate_create_application():
        """Симуляция создания заявки"""
        # Симуляция API вызова
        await asyncio.sleep(random.uniform(0.01, 0.05))
        
        # Создаем реальную заявку в Excel файле
        try:
            if ExcelIntegration is not None:
                excel = ExcelIntegration(data_dir="data")
                
                # Генерируем тестовые данные
                names = ["Тестовый", "Стресс", "Нагрузка", "Проверка", "Валидация"]
                surnames = ["Пользователь", "Клиент", "Заявитель", "Тестер", "Админ"]
                directions = ["Москва", "СПб", "Красноярск", "Сочи", "Новосибирск"]
                positions = ["Инженер", "Менеджер", "Аналитик", "Разработчик", "Тестировщик"]
                
                # Случайные данные
                name = f"{random.choice(names)} {random.choice(surnames)}"
                direction = random.choice(directions)
                position = random.choice(positions)
                tab_num = str(random.randint(10000, 99999))
                
                # Создаем заявку
                from datetime import datetime, timedelta
                now = datetime.now()
                departure_date = now + timedelta(days=random.randint(1, 30))
                
                application_data = {
                    'Время подачи заявки': now.strftime('%Y-%m-%d %H:%M:%S'),
                    'ФИО': name,
                    'Табельный номер': tab_num,
                    'Должность': position,
                    'Направление': direction,
                    'Информация о рейсе': f"{departure_date.strftime('%Y-%m-%d')} {random.randint(8, 20):02d}:{random.randint(0, 59):02d} SU-{random.randint(1000, 9999)}"
                }
                
                # Добавляем заявку в Excel
                excel.add_application(
                    oke='ОКЭ 1',
                    fio=application_data['ФИО'],
                    tab_num=application_data['Табельный номер'],
                    position=application_data['Должность'],
                    direction=application_data['Направление'],
                    flight_info=application_data['Информация о рейсе']
                )
                
        except Exception as e:
            print(f"Ошибка создания реальной заявки: {e}")
        
        # Симуляция случайных ошибок (5% вероятность)
        if random.random() < 0.05:
            raise Exception("Ошибка создания заявки")
    
    async def simulate_create_user():
        """Симуляция создания пользователя"""
        # Симуляция API вызова
        await asyncio.sleep(random.uniform(0.02, 0.08))
        
        # Симуляция случайных ошибок (3% вероятность)
        if random.random() < 0.03:
            raise Exception("Ошибка создания пользователя")
    
    async def simulate_send_notification():
        """Симуляция отправки уведомления"""
        # Симуляция API вызова
        await asyncio.sleep(random.uniform(0.005, 0.02))
        
        # Симуляция случайных ошибок (2% вероятность)
        if random.random() < 0.02:
            raise Exception("Ошибка отправки уведомления")
    
    async def run_concurrent_operations():
        """Запуск параллельных операций"""
        semaphore = asyncio.Semaphore(concurrency)
        
        async def limited_operation(operation_id):
            async with semaphore:
                await simulate_operation(operation_id)
        
        # Создаем задачи для всех операций
        tasks = []
        for i in range(count):
            if not stress_test_running:
                break
            task = asyncio.create_task(limited_operation(i))
            tasks.append(task)
        
        # Ждем завершения всех задач
        await asyncio.gather(*tasks, return_exceptions=True)
    
    try:
        # Запускаем асинхронный тест
        asyncio.run(run_concurrent_operations())
        
        # Завершаем тест
        stress_test_data['status'] = 'completed'
        stress_test_data['end_time'] = time.time()
        
    except Exception as e:
        stress_test_data['status'] = 'error'
        stress_test_data['error'] = str(e)
        stress_test_data['end_time'] = time.time()
    
    finally:
        stress_test_running = False

if __name__ == '__main__':
    print(f"🚀 Запуск Admin Panel Server...")
    print(f"📱 Bot ID: {config['bot_id']}")
    print(f"🌐 Server: http://{config['host']}:{config['port']}")
    print(f"👥 Admin Panel: http://{config['host']}:{config['port']}/admin")
    
    app.run(host=config['host'], port=config['port'], debug=True)
